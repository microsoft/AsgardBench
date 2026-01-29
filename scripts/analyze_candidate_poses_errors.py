#!/usr/bin/env python3
"""
Analyze pose errors by action_name and object_name.
Shows count, percent of total, and task_failed stats.

Also generates an Error Analysis sheet (like generate_reports) but only for
files that have the candidate_poses_errors property.
"""

import os
import re
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Any

import pandas as pd

import Magmathor.constants as c
from Magmathor.Utils.json_utils import FileReadError, read_json_file


def get_short_object_name(object_name: str) -> str:
    """Extract short name from object like 'SinkBasin_addc23' -> 'SinkBasin'"""
    if not object_name:
        return "(none)"
    # Strip "Types: " prefix if present
    if object_name.startswith("Types: "):
        object_name = object_name[7:]
    # Split on underscore and take everything before the last part if it looks like an ID
    parts = object_name.rsplit("_", 1)
    if len(parts) == 2 and len(parts[1]) >= 4:
        # Likely has an ID suffix
        return parts[0]
    return object_name


def find_test_files(base_path: str) -> list[str]:
    """Find all gpt-4o test_results.json files by constructing paths directly (faster than find on blobfuse).
    Skips folders that have both test_results.json and _test_results.json (indicates incomplete/rerun).
    """
    import os

    files = []
    skipped = 0
    for p in range(1, 7):
        benchmark_dir = f"{base_path}/magt_benchmark_p{p}"
        if not os.path.isdir(benchmark_dir):
            continue
        for dirname in os.listdir(benchmark_dir):
            if dirname.startswith("gpt-4o--"):
                test_file = f"{benchmark_dir}/{dirname}/test_results.json"
                backup_file = f"{benchmark_dir}/{dirname}/_test_results.json"
                if os.path.exists(test_file):
                    # Skip if there's also a _test_results.json file
                    if os.path.exists(backup_file):
                        skipped += 1
                        continue
                    files.append(test_file)
    if skipped > 0:
        print(f"Skipped {skipped} folders with _test_results.json backup files")
    return files


def process_single_file(filepath: str) -> dict:
    """Process a single test_results.json file and return stats.

    Returns all errors for Error Analysis, plus pose-specific stats.
    """
    file_stats = defaultdict(
        lambda: {
            "count": 0,
            "plan_count": 0,
            "failed_plan_count": 0,
            "example_path": None,
        }
    )
    total_errors = 0
    total_plans_with_errors = 0
    total_plans = 0
    failed_plans_with_pose_errors = 0
    # List of individual plan errors: [(action, object, failed, plan_path), ...]
    plan_errors = []
    # All step errors for Error Analysis (not just pose errors)
    all_step_errors = []

    try:
        data = read_json_file(filepath)
    except FileReadError as e:
        print(f"Skipping unreadable file: {e}")
        return {"skipped": True}

    test_results = data.get("test_results", [])
    dir_path = filepath.rsplit("/", 1)[0]

    for result in test_results:
        total_plans += 1
        task_failed = result.get("task_failed", False)
        step_errors = result.get("step_errors") or []
        task_name = result.get("task_name", "")

        # Track which (action, object) keys we've seen in this plan
        keys_seen_in_plan = set()
        plan_has_error = False
        # Get plan directory path
        plan_path = f"{dir_path}/Plans/{task_name}/"

        for error in step_errors:
            error_msg = error.get("error_msg", "")
            action_name = error.get("action_name", "(unknown)")

            # Collect ALL errors for Error Analysis
            if error_msg:
                all_step_errors.append(
                    {
                        "action_name": action_name,
                        "error_msg": error_msg,
                        "plan_path": plan_path,
                    }
                )

            # Check for pose errors specifically
            if c.POSES_ERROR in error_msg:
                plan_has_error = True
                object_name = error.get("object_name", "")
                short_name = get_short_object_name(object_name)

                key = (action_name, short_name)
                file_stats[key]["count"] += 1
                total_errors += 1

                # Count unique plans per key (and failed plans)
                if key not in keys_seen_in_plan:
                    keys_seen_in_plan.add(key)
                    file_stats[key]["plan_count"] += 1
                    if task_failed:
                        file_stats[key]["failed_plan_count"] += 1
                    # Store first example path
                    if file_stats[key]["example_path"] is None:
                        file_stats[key]["example_path"] = plan_path

                    # Add to individual plan errors list (one entry per unique action/object per plan)
                    plan_errors.append(
                        {
                            "action": action_name,
                            "object": short_name,
                            "failed": task_failed,
                            "plan_path": plan_path,
                        }
                    )

        if plan_has_error:
            total_plans_with_errors += 1
            if task_failed:
                failed_plans_with_pose_errors += 1

    return {
        "skipped": False,
        "stats": dict(file_stats),
        "total_errors": total_errors,
        "total_plans_with_errors": total_plans_with_errors,
        "total_plans": total_plans,
        "failed_plans_with_pose_errors": failed_plans_with_pose_errors,
        "plan_errors": plan_errors,
        "all_step_errors": all_step_errors,
    }


def analyze_files(base_path: str, limit: int = None):
    """Analyze all gpt-4o test_results.json files using parallel processing.

    Skips files with _test_results.json backup files.
    """

    # Structure: {(action_name, short_object_name): {"count": int, "plan_count": int, "failed_plan_count": int, "example_path": str}}
    error_stats = defaultdict(
        lambda: {
            "count": 0,
            "plan_count": 0,
            "failed_plan_count": 0,
            "example_path": None,
        }
    )
    total_errors = 0
    total_plans_with_errors = 0
    total_plans = 0
    failed_plans_with_pose_errors = 0
    all_plan_errors = []  # List of all individual plan errors
    all_step_errors = []  # All errors for Error Analysis
    files_processed = 0  # Count files successfully processed

    files = find_test_files(base_path)
    if limit:
        files = files[:limit]
    print(f"Found {len(files)} test_results.json files")

    completed = 0
    with ThreadPoolExecutor(max_workers=16) as executor:
        futures = {executor.submit(process_single_file, f): f for f in files}

        for future in as_completed(futures):
            completed += 1
            if completed % 20 == 0:
                print(f"Processing {completed}/{len(files)}...", flush=True)

            result = future.result()

            # Skip files that couldn't be read
            if result.get("skipped", False):
                continue

            files_processed += 1
            total_errors += result["total_errors"]
            total_plans_with_errors += result["total_plans_with_errors"]
            total_plans += result["total_plans"]
            failed_plans_with_pose_errors += result["failed_plans_with_pose_errors"]
            all_plan_errors.extend(result["plan_errors"])
            all_step_errors.extend(result.get("all_step_errors", []))

            # Merge stats
            for key, stats in result["stats"].items():
                error_stats[key]["count"] += stats["count"]
                error_stats[key]["plan_count"] += stats["plan_count"]
                error_stats[key]["failed_plan_count"] += stats["failed_plan_count"]
                # Keep first example path
                if error_stats[key]["example_path"] is None and stats["example_path"]:
                    error_stats[key]["example_path"] = stats["example_path"]

    print(f"Successfully processed {files_processed} files")

    return (
        error_stats,
        total_errors,
        total_plans_with_errors,
        total_plans,
        failed_plans_with_pose_errors,
        all_plan_errors,
        all_step_errors,
    )


def build_error_analysis(all_step_errors: list) -> tuple[list[str], list[list[Any]]]:
    """Build error analysis data from all step errors.

    Normalizes error messages by replacing object names with placeholders,
    then aggregates by normalized message.

    Returns:
        Tuple of (headers, rows) for the error analysis table.
    """
    # Words to exclude from object matching (common English words)
    excluded_words = {
        "Cannot",
        "The",
        "This",
        "That",
        "There",
        "Here",
        "What",
        "When",
        "Where",
        "Which",
        "While",
        "With",
        "Without",
        "Could",
        "Would",
        "Should",
        "Must",
        "Will",
        "Shall",
        "Have",
        "Has",
        "Had",
        "Does",
        "Did",
        "Done",
        "Been",
        "Being",
        "Was",
        "Were",
        "Are",
        "Is",
        "Am",
        "Be",
        "Not",
        "But",
        "And",
        "For",
        "From",
        "Into",
        "Onto",
        "Over",
        "Under",
        "Above",
        "Below",
        "Between",
        "Among",
        "Through",
        "During",
        "Before",
        "After",
        "Since",
        "Until",
        "Already",
        "Also",
        "Always",
        "Never",
        "Ever",
        "Only",
        "Just",
        "Still",
        "Even",
        "Very",
        "Too",
        "Much",
        "Many",
        "Some",
        "Any",
        "All",
        "Each",
        "Every",
        "Both",
        "Either",
        "Neither",
        "Other",
        "Another",
        "Such",
        "Same",
        "Different",
        "PUT",
        "GET",
        "Ran",
        "Failed",
        "Error",
        "Invalid",
        "Missing",
        "Expected",
        "Unexpected",
        "Unable",
        "Slice",
        "Sliced",
        "Types",
        "Type",
        "Action",
        "Object",
        "Objects",
        "Specifier",
        "Candidate",
        "Candidates",
        "None",
        "True",
        "False",
    }

    # Pattern to match object names with suffixes
    object_with_suffix_pattern = re.compile(
        r"[A-Z][a-zA-Z]*(?:_[a-zA-Z0-9]+)+(?:\(Clone\))?"
    )
    # Pattern to match simple CamelCase object names (e.g., WineBottle, SinkBasin)
    simple_object_pattern = re.compile(r"\b([A-Z][a-z]+(?:[A-Z][a-z]+)+)\b")

    def extract_object_type(obj_name: str) -> str:
        """Extract the base object type from a full object name."""
        name = obj_name.replace("(Clone)", "")
        parts = name.split("_")
        type_parts = []
        for part in parts:
            if re.match(r"^[a-f0-9]{6,}$", part, re.IGNORECASE):
                break
            if re.match(r"^\d+$", part):
                break
            type_parts.append(part)
        return "_".join(type_parts) if type_parts else parts[0]

    def find_objects_in_message(msg: str) -> list:
        """Find all object references in an error message."""
        result = []
        # First find objects with suffixes (e.g., Bread_fe4bb3e3)
        for match in object_with_suffix_pattern.findall(msg):
            if match not in result:
                result.append(match)
        # Find simple CamelCase objects (e.g., WineBottle, SinkBasin)
        for match in simple_object_pattern.findall(msg):
            if match not in excluded_words and match not in result:
                # Check it's not already covered by a suffixed version
                if not any(m.startswith(match + "_") for m in result):
                    result.append(match)
        # Also find single capitalized words that might be objects
        single_word_pattern = re.compile(r"\b([A-Z][a-z]{2,})\b")
        for match in single_word_pattern.findall(msg):
            if match not in excluded_words and match not in result:
                if not any(m.startswith(match + "_") for m in result):
                    result.append(match)
        return result

    # Structure to track error info
    @dataclass
    class ErrorInfo:
        count: int = 0
        placeholder_values: dict = field(default_factory=dict)
        example_plan: str = ""  # First plan that had this error

    # Aggregate errors: {full_error: ErrorInfo}
    error_data: dict[str, ErrorInfo] = defaultdict(ErrorInfo)
    total_error_count = 0

    for error in all_step_errors:
        action_name = error.get("action_name", "")
        error_msg = error.get("error_msg", "")
        plan_path = error.get("plan_path", "")
        if not error_msg:
            continue

        total_error_count += 1

        # Find all unique object names in the error message
        matches = find_objects_in_message(error_msg)
        unique_objects = []
        for m in matches:
            if m not in unique_objects:
                unique_objects.append(m)

        # Replace each unique object with {X}, {Y}, {Z}, etc.
        placeholder_names = ["X", "Y", "Z", "W", "V"]
        normalized_msg = error_msg
        object_types = {}  # placeholder_name -> object_type
        for i, obj_name in enumerate(unique_objects):
            ph_name = placeholder_names[i] if i < len(placeholder_names) else f"O{i}"
            placeholder = "{" + ph_name + "}"
            normalized_msg = normalized_msg.replace(obj_name, placeholder)
            object_types[ph_name] = extract_object_type(obj_name)

        # Format: "action_name : normalized_error_msg"
        full_error = (
            f"{action_name} : {normalized_msg}" if action_name else normalized_msg
        )

        # Update error info
        error_info = error_data[full_error]
        error_info.count += 1

        # Store first example plan path
        if not error_info.example_plan and plan_path:
            error_info.example_plan = plan_path

        # Track placeholder values
        for ph_name, obj_type in object_types.items():
            if ph_name not in error_info.placeholder_values:
                error_info.placeholder_values[ph_name] = {}
            error_info.placeholder_values[ph_name][obj_type] = (
                error_info.placeholder_values[ph_name].get(obj_type, 0) + 1
            )

    if not error_data:
        return [], []

    # Helper to format placeholder values
    def format_placeholder_values(ph_dict: dict) -> str:
        if not ph_dict:
            return ""
        total = sum(ph_dict.values())
        sorted_items = sorted(ph_dict.items(), key=lambda x: x[1], reverse=True)
        parts = [
            f"{obj_type} ({count/total*100:.0f}%)" for obj_type, count in sorted_items
        ]
        return ", ".join(parts)

    # Build output data
    headers = ["pct", "action", "error_message", "X", "Y", "Z", "example_plan"]
    rows = []

    # Sort errors by count descending
    sorted_errors = sorted(error_data.items(), key=lambda x: x[1].count, reverse=True)

    for full_error, error_info in sorted_errors:
        pct = error_info.count / total_error_count if total_error_count > 0 else 0

        # Parse "action : message" format
        if " : " in full_error:
            action, message = full_error.split(" : ", 1)
        else:
            action = ""
            message = full_error

        # Get placeholder values
        x_values = format_placeholder_values(error_info.placeholder_values.get("X", {}))
        y_values = format_placeholder_values(error_info.placeholder_values.get("Y", {}))
        z_values = format_placeholder_values(error_info.placeholder_values.get("Z", {}))

        rows.append(
            [
                pct,
                action,
                message,
                x_values,
                y_values,
                z_values,
                error_info.example_plan,
            ]
        )

    return headers, rows


def main():
    import sys

    # Quick test mode: pass --test or -t to only process 3 files
    limit = 3 if any(arg in sys.argv for arg in ["--test", "-t"]) else None

    base_path = "/mnt/magmathor/20260115_Test"

    print(f"Analyzing candidate poses errors in {base_path}...")
    print()

    (
        error_stats,
        total_errors,
        total_plans_with_errors,
        total_plans,
        failed_plans_with_pose_errors,
        all_plan_errors,
        all_step_errors,
    ) = analyze_files(base_path, limit=limit)

    if total_plans == 0:
        print("No test files were successfully processed.")
        return

    # Print summary table first
    print()
    print("=" * 60)
    print(f"SUMMARY: Plans with '{c.POSES_ERROR}' errors")
    print("=" * 60)
    print(f"{'Total plans processed:':<45} {total_plans:>10}")
    print(f"{'Plans with pose errors:':<45} {total_plans_with_errors:>10}")
    print(f"{'Failed plans with pose errors:':<45} {failed_plans_with_pose_errors:>10}")
    pct_failed = (
        (failed_plans_with_pose_errors / total_plans * 100) if total_plans > 0 else 0
    )
    print(f"{'% of all plans that failed with pose errors:':<45} {pct_failed:>9.1f}%")
    print("=" * 60)
    print()

    # Sort by count descending
    sorted_stats = sorted(
        error_stats.items(), key=lambda x: x[1]["count"], reverse=True
    )

    if total_errors > 0:
        print()
        print("DETAILS: Errors by Action/Object")
        print()
        # Print header
        print(
            f"{'Action':<15} {'Object':<25} {'Count':>6} {'Frac':>6}   {'Plans':>6}   {'Failed':>10}   Example"
        )
        print("-" * 85)

        for (action_name, object_name), stats in sorted_stats:
            count = stats["count"]
            plan_count = stats["plan_count"]
            failed_plan_count = stats["failed_plan_count"]
            example_path = stats["example_path"] or ""

            frac_of_total = count / total_errors
            failed_fraction = f"{failed_plan_count}/{plan_count}"

            print(
                f"{action_name:<15} {object_name:<25} {count:>6} {frac_of_total:>6.2f}   {plan_count:>6}   {failed_fraction:>10}   {example_path}"
            )

        print("-" * 85)
        print(
            f"{'TOTAL':<42} {total_errors:>6} {'1.00':>6}   {total_plans_with_errors:>6}"
        )
    else:
        print(f"No '{c.POSES_ERROR}' errors found (generating Error Analysis only).")

    # Create Excel file
    rows = []
    for (action_name, object_name), stats in sorted_stats:
        count = stats["count"]
        plan_count = stats["plan_count"]
        failed_plan_count = stats["failed_plan_count"]
        example_path = stats["example_path"] or ""

        frac_of_total = count / total_errors if total_errors > 0 else 0
        failed_frac = failed_plan_count / plan_count if plan_count > 0 else 0

        rows.append(
            {
                "Action": action_name,
                "Object": object_name,
                "Count": count,
                "Frac": frac_of_total,
                "Plans": plan_count,
                "Failed": failed_plan_count,
                "Failed Frac": failed_frac,
                "Example": example_path,
            }
        )

    # Add total row if there are pose errors
    if total_errors > 0:
        rows.append(
            {
                "Action": "TOTAL",
                "Object": "",
                "Count": total_errors,
                "Frac": 1.0,
                "Plans": total_plans_with_errors,
                "Failed": "",
                "Failed Frac": "",
                "Example": "",
            }
        )

    df = pd.DataFrame(rows)
    output_path = "/home/larsliden/Magmathor/Generated/candidate_poses_errors.xlsx"

    # Create summary dataframe
    pct_failed = failed_plans_with_pose_errors / total_plans if total_plans > 0 else 0
    summary_rows = [
        {"Metric": "Total plans processed", "Value": total_plans},
        {"Metric": "Plans with pose errors", "Value": total_plans_with_errors},
        {
            "Metric": "Failed plans with pose errors",
            "Value": failed_plans_with_pose_errors,
        },
        {"Metric": "% of all plans that failed with pose errors", "Value": pct_failed},
    ]
    summary_df = pd.DataFrame(summary_rows)

    def autosize_columns(worksheet):
        """Autosize all columns in a worksheet based on content."""
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Add a little padding
            adjusted_width = min(
                max_length + 2, 100
            )  # Cap at 100 to avoid very wide columns
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def add_table(worksheet, df, table_name):
        """Add a table to a worksheet for filtering/sorting."""
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # Calculate table range (accounting for index=False, so data starts at row 1)
        max_row = len(df) + 1  # +1 for header
        max_col = len(df.columns)
        table_range = f"A1:{get_column_letter(max_col)}{max_row}"

        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)

    # Write to Excel with percentage formatting
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write summary sheet first
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        summary_ws = writer.sheets["Summary"]
        # Format the percentage row
        summary_ws.cell(row=5, column=2).number_format = "0.0%"
        autosize_columns(summary_ws)

        # Write details sheet (pose errors only)
        df.to_excel(writer, index=False, sheet_name="Pose Errors")
        worksheet = writer.sheets["Pose Errors"]

        # Format Frac column (D) and Failed Frac column (G) as percentages
        from openpyxl.styles import numbers

        for row in range(2, len(df) + 2):  # Skip header row
            worksheet.cell(row=row, column=4).number_format = "0%"  # Frac column
            cell = worksheet.cell(row=row, column=7)  # Failed Frac column
            if cell.value != "":
                cell.number_format = "0%"
        autosize_columns(worksheet)
        if len(df) > 0:
            add_table(worksheet, df, "PoseErrorsTable")

        # Write individual plan errors sheet (pose errors only)
        # Remove duplicates (same action/object/failed/plan combination)
        seen = set()
        unique_plan_errors = []
        for e in all_plan_errors:
            key = (e["action"], e["object"], e["failed"], e["plan_path"])
            if key not in seen:
                seen.add(key)
                unique_plan_errors.append(e)

        # Sort by action, then object, then failed status
        sorted_plan_errors = sorted(
            unique_plan_errors,
            key=lambda x: (x["action"], x["object"], not x["failed"]),
        )
        plan_errors_df = pd.DataFrame(
            [
                {
                    "Action": e["action"],
                    "Object": e["object"],
                    "Failed": "Yes" if e["failed"] else "No",
                    "Plan": e["plan_path"]
                    .rstrip("/")
                    .rsplit("/", 1)[-1],  # Extract plan name from path
                    "Plan Directory": e["plan_path"],
                }
                for e in sorted_plan_errors
            ]
        )
        plan_errors_df.to_excel(writer, index=False, sheet_name="All Plan Errors")
        plan_errors_ws = writer.sheets["All Plan Errors"]
        autosize_columns(plan_errors_ws)
        if len(plan_errors_df) > 0:
            add_table(plan_errors_ws, plan_errors_df, "AllPlanErrorsTable")

        # Write Error Analysis sheet (all errors, not just pose errors)
        error_headers, error_rows = build_error_analysis(all_step_errors)
        if error_rows:
            error_analysis_df = pd.DataFrame(error_rows, columns=error_headers)
            error_analysis_df.to_excel(writer, index=False, sheet_name="Error Analysis")
            error_ws = writer.sheets["Error Analysis"]

            # Format pct column (A) as percentage
            for row in range(2, len(error_analysis_df) + 2):
                error_ws.cell(row=row, column=1).number_format = "0.0%"

            autosize_columns(error_ws)
            add_table(error_ws, error_analysis_df, "ErrorAnalysisTable")
            print(
                f"Error Analysis: {len(error_rows)} unique error patterns from {len(all_step_errors)} total errors"
            )

    print()
    print(f"Excel file saved to: {output_path}")


if __name__ == "__main__":
    main()
